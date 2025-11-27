from io import BytesIO
from sqlalchemy.orm import Session
from models.medicine import Medicine
from schemas.medicine_schema import MedicineCreate, MedicineUpdate
import openpyxl

# ========== CRUD ==========

def get_medicines(db: Session, skip: int = 0, limit: int = 100):
    return db.query(Medicine).offset(skip).limit(limit).all()


def get_medicine(db: Session, medicine_id: int):
    return db.query(Medicine).filter(Medicine.id == medicine_id).first()


def create_medicine(db: Session, medicine: MedicineCreate):
    db_medicine = Medicine(**medicine.dict(exclude_unset=True))
    db.add(db_medicine)
    db.commit()
    db.refresh(db_medicine)
    return db_medicine


def update_medicine(db: Session, medicine_id: int, medicine: MedicineUpdate):
    db_medicine = get_medicine(db, medicine_id)
    if not db_medicine:
        return None

    for field, value in medicine.dict(exclude_unset=True).items():
        setattr(db_medicine, field, value)

    db.commit()
    db.refresh(db_medicine)
    return db_medicine


def delete_medicine(db: Session, medicine_id: int):
    db_medicine = get_medicine(db, medicine_id)
    if not db_medicine:
        return None

    db.delete(db_medicine)
    db.commit()
    return db_medicine


# ========== DOWNLOAD INVENTORY ==========

def get_medicine_inventory_excel(db: Session) -> BytesIO:
    medicines = db.query(Medicine).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medicine Inventory"

    # Headers
    ws.append([
        "ID", "Name", "Brand", "Category","Quantity", "Expiry Date",
        "Cost", "Tax", "Total Cost"
    ])

    for med in medicines:
        ws.append([
            med.id,
            med.name,
            med.brand,
            med.category,
            med.quantity,
            str(med.expiry_date) if med.expiry_date else None,
            med.cost,
            med.tax,
            med.total_cost
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ========== IMPORT MEDICINE EXCEL ==========

def import_medicine_inventory_excel(file, db: Session):
    try:
        workbook = openpyxl.load_workbook(BytesIO(file.file.read()))
        sheet = workbook.active

        inserted, updated = 0, 0

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if not any(row):
                continue

            name = str(row[0]).strip().upper() if row[0] else None
            brand = str(row[1]).strip() if row[1] else None
            category = str(row[2]).strip() if row[2] else None
            quantity = int(row[3]) if row[3] else 0
            expiry_date = row[4] if row[4] else None
            cost = float(row[5]) if row[5] else 0
            tax = float(row[6]) if row[6] else 0
            total_cost = float(row[7]) if row[7] else (cost + tax)

            if not name:
                continue

            existing = db.query(Medicine).filter(Medicine.name == name).first()

            if existing:
                existing.brand = brand or existing.brand
                existing.category = category or existing.category
                existing.quantity = quantity
                existing.expiry_date = expiry_date or existing.expiry_date
                existing.cost = cost
                existing.tax = tax
                existing.total_cost = total_cost
                updated += 1
            else:
                new_medicine = Medicine(
                    name=name,
                    brand=brand,
                    quantity=quantity,
                    cost=cost,
                    tax=tax,
                    total_cost=total_cost,
                    category=category,
                    expiry_date=expiry_date
                )
                db.add(new_medicine)
                inserted += 1

        db.commit()
        return {"message": "Import completed successfully", "inserted": inserted, "updated": updated}

    except Exception as e:
        db.rollback()
        return {"error": f"Failed to import Excel file: {str(e)}"}


# ========== INDENT APPROVAL ==========

def approve_indent_excel(file, db: Session):
    from models.medicine import Medicine

    contents = file.file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    inserted = 0
    updated = 0

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        try:
            s_no, drug_name, brand, present_qty, required_qty, cost, tax, total_cost = row[:8]

            name = str(drug_name).strip().upper() if drug_name else None
            brand = str(brand).strip() if brand else None

            if not name:
                continue

            existing = db.query(Medicine).filter(Medicine.name == name).first()

            if existing:
                existing.quantity = (existing.quantity or 0) + int(required_qty or 0)
                existing.brand = brand or existing.brand
                existing.cost = cost or existing.cost
                existing.tax = tax or existing.tax
                existing.total_cost = total_cost or existing.total_cost
                updated += 1
            else:
                new_medicine = Medicine(
                    name=name,
                    brand=brand,
                    quantity=int(present_qty or 0) + int(required_qty or 0),
                    cost=cost,
                    tax=tax,
                    total_cost=total_cost,
                    category=None,
                    expiry_date=None
                )
                db.add(new_medicine)
                inserted += 1
        except Exception:
            continue

    db.commit()
    return {"inserted": inserted, "updated": updated}